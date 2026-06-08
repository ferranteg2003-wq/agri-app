import streamlit as st
import cv2
import numpy as np
from PIL import Image
from pillow_heif import register_heif_opener
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# Setup pagina
st.set_page_config(page_title="AgriApp - Test in Cloud", layout="centered")
register_heif_opener()

def analizza_cartina_da_upload(uploaded_file):
    image_bytes = uploaded_file.read()
    
    try:
        if uploaded_file.name.lower().endswith('.heic'):
            pil_image = Image.open(io.BytesIO(image_bytes))
            numpy_image = np.array(pil_image)
            img = cv2.cvtColor(numpy_image, cv2.COLOR_RGB2BGR)
        else:
            nparr = np.frombuffer(image_bytes, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None: return None, "Errore immagine."
    except Exception as e: return None, f"Errore: {str(e)}"

    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    mask_yellow = cv2.inRange(hsv, np.array([10, 40, 40]), np.array([50, 255, 255]))
    mask_blue = cv2.inRange(hsv, np.array([90, 30, 30]), np.array([160, 255, 255]))
    mask_dark = cv2.inRange(hsv, np.array([0, 0, 0]), np.array([180, 255, 90]))

    mask_cartina_fisica = cv2.bitwise_or(mask_yellow, mask_blue)
    mask_cartina_fisica = cv2.bitwise_or(mask_cartina_fisica, mask_dark)
    mask_cartina_fisica = cv2.morphologyEx(mask_cartina_fisica, cv2.MORPH_CLOSE, np.ones((21, 21), np.uint8))

    contours, _ = cv2.findContours(mask_cartina_fisica, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours: return None, "Nessuna cartina."

    x, y, w, h = cv2.boundingRect(max(contours, key=cv2.contourArea))
    margine = 40
    y1 = max(0, y - margine); y2 = min(img.shape[0], y + h + margine)
    x1 = max(0, x - margine); x2 = min(img.shape[1], x + w + margine)
    
    img_con_margine = img[y1:y2, x1:x2].copy()
    hsv_con_margine = hsv[y1:y2, x1:x2].copy()

    mask_forma_esatta = np.zeros((y2-y1, x2-x1), dtype=np.uint8)
    cv2.drawContours(mask_forma_esatta, [max(contours, key=cv2.contourArea) - [x1, y1]], -1, 255, thickness=cv2.FILLED)

    mask_analisi_interna = cv2.erode(mask_forma_esatta, np.ones((7,7), np.uint8), iterations=5)
    mask_giallo_pulito = cv2.inRange(hsv_con_margine, np.array([12, 60, 110]), np.array([50, 255, 255]))
    mask_bagnata = cv2.bitwise_and(cv2.bitwise_not(mask_giallo_pulito), cv2.bitwise_not(mask_giallo_pulito), mask=mask_analisi_interna)

    pix_orig = max(1, cv2.countNonZero(mask_forma_esatta))
    pix_analizzati = max(1, cv2.countNonZero(mask_analisi_interna))
    area_reale = (pix_analizzati / pix_orig) * 19.76
    
    pix_bagnati = cv2.countNonZero(mask_bagnata)
    perc_bagnata = (pix_bagnati / pix_analizzati) * 100
    area_bagnata = (perc_bagnata / 100) * area_reale

    num_labels, _, stats, _ = cv2.connectedComponentsWithStats(mask_bagnata, connectivity=8)
    num_gocce = max(0, num_labels - 1)
    
    mm2_px = 1976.0 / pix_orig
    classi = {"micro": 0, "piccole": 0, "medie": 0, "grandi": 0, "extra": 0}
    for i in range(1, num_labels):
        area_mm2 = stats[i, cv2.CC_STAT_AREA] * mm2_px
        if area_mm2 < 0.1: classi["micro"] += 1
        elif area_mm2 < 0.5: classi["piccole"] += 1
        elif area_mm2 < 1.0: classi["medie"] += 1
        elif area_mm2 < 2.0: classi["grandi"] += 1
        else: classi["extra"] += 1

    img_rosse = img_con_margine.copy()
    img_rosse[mask_bagnata > 0] = [0, 0, 255]

    dati = {
        "file": uploaded_file.name, "bagnatura": perc_bagnata, "area_reale": area_reale,
        "area_bagnata": area_bagnata, "gocce": num_gocce, "densita": num_gocce / area_reale,
        "pulita": 100 - perc_bagnata, "distribuzione": classi
    }
    return (cv2.cvtColor(img_rosse, cv2.COLOR_BGR2RGB), dati), None

def genera_excel(risultati):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisi"
    headers = ["Nome File", "Area Reale (cm²)", "Bagnatura (%)", "N° Gocce", "Densità", "Micro", "Piccole", "Medie", "Grandi", "Extra"]
    ws.append(headers)
    
    for r, res in enumerate(risultati, 2):
        ws.cell(r, 1, res['file'])
        ws.cell(r, 2, res['area_reale']).number_format = '0.00'
        ws.cell(r, 3, res['bagnatura'] / 100).number_format = '0.00%'
        ws.cell(r, 4, res['gocce'])
        ws.cell(r, 5, res['densita']).number_format = '0.0'
        ws.cell(r, 6, res['distribuzione']['micro'])
        ws.cell(r, 7, res['distribuzione']['piccole'])
        ws.cell(r, 8, res['distribuzione']['medie'])
        ws.cell(r, 9, res['distribuzione']['grandi'])
        ws.cell(r, 10, res['distribuzione']['extra'])
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

st.title("💧 AgriApp - Analisi Cartine (Cloud)")
if 'risultati' not in st.session_state: st.session_state.risultati = []

files = st.file_uploader("Scatta o seleziona foto (JPG, PNG, HEIC)", type=['jpg', 'jpeg', 'png', 'heic'], accept_multiple_files=True)

if files and st.button("🚀 Analizza"):
    st.session_state.risultati = []
    for f in files:
        res, err = analizza_cartina_da_upload(f)
        if err: st.error(err)
        else:
            img, dati = res
            st.session_state.risultati.append(dati)
            col1, col2 = st.columns([1, 1])
            col1.image(img, use_container_width=True)
            with col2:
                st.success(f"Bagnatura: {dati['bagnatura']:.1f}%")
                st.info(f"Densità: {dati['densita']:.1f} g/cm²")
                st.write(f"Gocce: {dati['gocce']}")

if st.session_state.risultati:
    st.download_button("📥 Scarica File Excel", genera_excel(st.session_state.risultati), "Analisi_Cartine.xlsx")